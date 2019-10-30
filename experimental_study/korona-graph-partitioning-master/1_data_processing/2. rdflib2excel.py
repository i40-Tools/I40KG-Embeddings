from openpyxl import Workbook
import re

rownumber=1 #Row number of the excel file where each row links to one paper/entity (each subject of the NT file).
visitedsubjects = [] #List of subjects already visited
InputNTfile = "output/ISWC.nt" #Input NT file
XlsxPath = "output/papers_authors.xlsx" #Excel File to which we have for output.

wb = Workbook() #open workbook
ws = wb.active # select the active worksheet# select the active worksheet

#Set up Header Row. Contents of Row 1 of the Excel File.
ws.cell(row=1, column=1, value= 'Paper Number')
ws.cell(row=1, column=2, value= 'Title')
ws.cell(row=1, column=3, value= 'Number of Authors')
ws.cell(row=1, column=4, value= 'Year')

def checkpredicate(predicate): #Check if we need the predicate and if we do then return to which column we should save it.
    if(predicate == "http://dblp.org/rdf/schema-2017-04-18#title"):
        return 2
    if (predicate =="http://dblp.org/rdf/schema-2017-04-18#authoredBy"):
        return 3
    if (predicate =="http://dblp.org/rdf/schema-2017-04-18#yearOfPublication"):
        return 4
    return 10

def checksubject(subject): #All triples for the same paper will have the same subject and belong in the same row. If subject changes we go to the next row.
    global rownumber
    global visitedsubjects
    if (subject not in visitedsubjects): # A new paper
        if (ws.cell(row = rownumber, column= 3).value >0):  #If the current row has 0 authors then we overwrite in that row since it cannot be used for processing.
            rownumber+=1                    #Go to the next row on the Excel File
        visitedsubjects.append(subject) #Append the current subject to the list of visited subjects.
        ws.cell(row=rownumber, column=1, value=rownumber-1) #Column 1 of excel file contains conf. paper number
        ws.cell(row=rownumber, column=3, value=0)           #Column 3 of excel file is number of authors. Set first to zero.
    return

                                                    ####***~~~~Execution starts from here~~~****#####

with open(InputNTfile) as f: #open input NT file.
        for line in f: #Iterate through the NT file line by line
            eachnumber = re.split(r'\t+', line)
            eachnumber[0]=eachnumber[0].strip("<>.")    #Subject
            eachnumber[1]=eachnumber[1].strip("<>.")    #Predicate
            eachnumber[2]=eachnumber[2]                 #Object
            checksubject(eachnumber[0]) #Set the row number to write on the Excel file depending on the subject.
            columnnumber = checkpredicate(eachnumber[1]) #Get column number of the excel file where to be saved for the predicate of the current triple.

            if (columnnumber == 3): #Number of authors column. We add the names of the authors at the right end columns 7 onwards
                temp = ws.cell(row = rownumber, column= 3).value #Get the current "number of author" column value of the excel file to know which column to write the author's name.
                authorcol=temp+5 #Get the free column number of the excel file to write the name of author.
                ws.cell(row=rownumber, column=authorcol, value=eachnumber[2].replace('<','').replace('>','')) #Write the name of author.
                ws.cell(row=rownumber, column=3, value=temp+1) #Update the "number of authors" column of the excel file.

            if (columnnumber<10 and columnnumber!=3): #Predicate is to be considered but value is not "author"
                if (columnnumber == 4):
                    ws.cell(row=rownumber,column=columnnumber,value=int(eachnumber[2].replace('"', ''))) #Save the predicate to the particular cell of the excel file.
                else:
                    ws.cell(row=rownumber,column=columnnumber,value=eachnumber[2].replace('"', '')) #Save the predicate to the particular cell of the excel file.
        f.close()
        wb.save(XlsxPath)
